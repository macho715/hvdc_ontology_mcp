from __future__ import annotations

import contextlib
import contextvars
from datetime import datetime, UTC
import json
import logging
import os
from pathlib import Path
import shutil
import sys
import threading
from typing import Any
import uuid

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from mcp.server.fastmcp import FastMCP
from starlette.requests import Request
from starlette.responses import JSONResponse

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

_correlation_id_var: contextvars.ContextVar[str] = contextvars.ContextVar(
    "excel_mcp_correlation_id",
    default="",
)
_workbook_locks_guard = threading.Lock()
_workbook_locks: dict[str, threading.Lock] = {}


class JsonLineFormatter(logging.Formatter):
    def format(self, record: logging.LogRecord) -> str:
        payload: dict[str, Any] = {
            "ts": datetime.now(UTC).isoformat(),
            "level": record.levelname,
            "logger": record.name,
            "message": record.getMessage(),
            "correlation_id": getattr(record, "correlation_id", None) or _correlation_id_var.get() or None,
        }
        for field in ("event", "path", "method", "status_code", "tool_name"):
            value = getattr(record, field, None)
            if value is not None:
                payload[field] = value
        if record.exc_info:
            payload["exception"] = self.formatException(record.exc_info)
        return json.dumps(payload, ensure_ascii=False)


_root_handler = logging.StreamHandler(stream=sys.stderr)
_root_handler.setFormatter(JsonLineFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_root_handler], force=True)
logger = logging.getLogger("excel_mcp")


def _resolve_root() -> Path:
    raw_root = os.getenv("EXCEL_MCP_ROOT", "workbooks")
    root = Path(raw_root)
    if not root.is_absolute():
        root = BASE_DIR / root
    resolved = root.resolve()
    resolved.mkdir(parents=True, exist_ok=True)
    return resolved


def _read_port() -> int:
    raw_port = os.getenv("EXCEL_MCP_PORT", "8002")
    try:
        port = int(raw_port)
    except ValueError as exc:
        raise ValueError(f"Invalid EXCEL_MCP_PORT: {raw_port}") from exc

    if port < 1 or port > 65535:
        raise ValueError(f"EXCEL_MCP_PORT out of range: {port}")

    return port


ROOT = _resolve_root()
HOST = os.getenv("EXCEL_MCP_HOST", "127.0.0.1")
PORT = _read_port()
EXPORT_VERSION = "2026-03-17"
DEFAULT_EXPORT_SHEETS = ("Summary", "ZERO", "KPI", "Exceptions")

mcp = FastMCP(
    name="Excel MCP Starter",
    host=HOST,
    port=PORT,
    streamable_http_path="/mcp",
    json_response=True,
    stateless_http=True,
)


def _current_correlation_id() -> str:
    return _correlation_id_var.get() or ""


def _new_correlation_id() -> str:
    return uuid.uuid4().hex


@contextlib.contextmanager
def _correlation_scope(correlation_id: str | None = None):
    token = _correlation_id_var.set(correlation_id or _new_correlation_id())
    try:
        yield _correlation_id_var.get()
    finally:
        _correlation_id_var.reset(token)


def _request_correlation_id(request: Request | None = None) -> str:
    if request is not None:
        for header in ("x-correlation-id", "x-request-id"):
            value = str(request.headers.get(header, "")).strip()
            if value:
                return value
    return _current_correlation_id() or _new_correlation_id()


def _log_event(level: int, message: str, **fields: Any) -> None:
    logger.log(level, message, extra=fields)


def _status_payload() -> dict[str, Any]:
    return {
        "name": "Excel MCP Starter",
        "transport": "streamable-http",
        "host": HOST,
        "port": PORT,
        "root": str(ROOT),
        "mcp_path": "/mcp",
        "tools": [
            "ping",
            "create_workbook",
            "list_workbooks",
            "get_workbook_info",
            "list_sheets",
            "read_sheet",
            "read_range",
            "append_rows",
            "write_cells",
            "write_range",
            "sheet_dimensions",
            "apply_hvdc_export",
        ],
        "export_contract_version": EXPORT_VERSION,
        "sheet_defaults": list(DEFAULT_EXPORT_SHEETS),
    }


def _safe_path(relative_path: str) -> Path:
    if not relative_path:
        raise ValueError("path is required")

    candidate = (ROOT / relative_path).resolve()
    if ROOT not in candidate.parents:
        raise ValueError("path escapes EXCEL_MCP_ROOT")

    if candidate.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("only .xlsx/.xlsm files are allowed")

    return candidate


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _relative_path(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def _workbook_lock(path: Path) -> threading.Lock:
    key = str(path)
    with _workbook_locks_guard:
        lock = _workbook_locks.get(key)
        if lock is None:
            lock = threading.Lock()
            _workbook_locks[key] = lock
        return lock


@contextlib.contextmanager
def _locked_workbook(path: Path, *, timeout_sec: float = 5.0):
    lock = _workbook_lock(path)
    acquired = lock.acquire(timeout=timeout_sec)
    if not acquired:
        raise TimeoutError(f"workbook is busy: {path.name}")
    try:
        yield
    finally:
        lock.release()


def _backup_workbook(path: Path) -> Path | None:
    if not path.exists():
        return None
    backup_path = path.with_suffix(path.suffix + ".bak")
    shutil.copy2(path, backup_path)
    return backup_path


def _workbook_metadata(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = load_workbook(path, data_only=False, read_only=True)
    sheet_rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_column = ws.max_column or 0
        sheet_rows.append(
            {
                "name": sheet_name,
                "max_row": ws.max_row or 0,
                "max_column": max_column,
                "last_column_letter": get_column_letter(max_column) if max_column else None,
            }
        )
    return {
        "ok": True,
        "path": str(path),
        "relative_path": _relative_path(path),
        "size_bytes": path.stat().st_size,
        "modified_at": datetime.fromtimestamp(path.stat().st_mtime, UTC).isoformat(),
        "sheet_count": len(wb.sheetnames),
        "sheets": sheet_rows,
    }


def _get_sheet(workbook_path: Path, sheet: str):
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=False)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {sheet}")

    return wb, wb[sheet]


def _headers_from_row(values: tuple[Any, ...] | list[Any] | None) -> list[str]:
    if not values:
        return []
    headers = []
    for index, value in enumerate(values, start=1):
        text = str(value).strip() if value not in (None, "") else f"column_{index}"
        headers.append(text)
    return headers


def _normalize_cell_value(value: Any) -> Any:
    return "" if value is None else value


def _serializable_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _sheet_records(
    path: Path,
    sheet: str,
    *,
    offset: int,
    limit: int,
    columns: list[str] | None,
) -> tuple[list[dict[str, Any]], list[str], int]:
    wb = load_workbook(path, data_only=False, read_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {sheet}")
    ws = wb[sheet]
    iterator = ws.iter_rows(values_only=True)
    header_values = next(iterator, None)
    headers = _headers_from_row(header_values)
    if not headers:
        return [], [], 0

    selected_columns = columns or headers
    missing_columns = [column for column in selected_columns if column not in headers]
    if missing_columns:
        raise KeyError(f"Columns not found: {', '.join(missing_columns)}")
    selected_indexes = [headers.index(column) for column in selected_columns]
    records: list[dict[str, Any]] = []
    total_rows = max((ws.max_row or 1) - 1, 0)
    for row_index, row in enumerate(iterator):
        if row_index < offset:
            continue
        if len(records) >= limit:
            break
        record = {
            column: _normalize_cell_value(row[index] if index < len(row) else None)
            for column, index in zip(selected_columns, selected_indexes)
        }
        records.append(record)
    return records, selected_columns, total_rows


def _row_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(str(key))
    return headers


def _write_row_dicts(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    headers = _row_headers(rows)
    for column_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=column_index, value=header)
    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(headers, start=1):
            ws.cell(
                row=row_index,
                column=column_index,
                value=_serializable_cell_value(row.get(header)),
            )


@mcp.custom_route("/", methods=["GET"], include_in_schema=False)
async def root_status(request: Request) -> JSONResponse:
    with _correlation_scope(_request_correlation_id(request)) as correlation_id:
        payload = _status_payload()
        payload["correlation_id"] = correlation_id
        _log_event(
            logging.INFO,
            "excel root status requested",
            event="root_status",
            path="/",
            method="GET",
            status_code=200,
            correlation_id=correlation_id,
        )
        return JSONResponse(payload)


@mcp.custom_route("/health", methods=["GET"], include_in_schema=False)
async def healthcheck(request: Request) -> JSONResponse:
    with _correlation_scope(_request_correlation_id(request)) as correlation_id:
        status = _status_payload()
        files = list(ROOT.rglob("*.xlsx")) + list(ROOT.rglob("*.xlsm"))
        payload = {
            **status,
            "status": "healthy",
            "checks": [
                {"name": "root_exists", "ok": ROOT.exists(), "detail": str(ROOT)},
                {"name": "tools_ready", "ok": len(status["tools"]) >= 8, "detail": f"tools={len(status['tools'])}"},
            ],
            "workbook_count": len(files),
            "correlation_id": correlation_id,
        }
        _log_event(
            logging.INFO,
            "excel health requested",
            event="healthcheck",
            path="/health",
            method="GET",
            status_code=200,
            correlation_id=correlation_id,
        )
        return JSONResponse(payload)


@mcp.tool()
def ping() -> dict[str, Any]:
    """Return a basic health payload."""
    return {
        "ok": True,
        "root": str(ROOT),
        "host": HOST,
        "port": PORT,
        "export_contract_version": EXPORT_VERSION,
    }


@mcp.tool()
def create_workbook(path: str, sheets: list[str] | None = None) -> dict[str, Any]:
    """Create or overwrite a workbook under EXCEL_MCP_ROOT."""
    target = _safe_path(path)
    _ensure_parent(target)
    with _correlation_scope():
        with _locked_workbook(target):
            backup_path = _backup_workbook(target)
            wb = Workbook()
            if sheets:
                default = wb.active
                default.title = sheets[0]
                for name in sheets[1:]:
                    wb.create_sheet(title=name)
            wb.save(target)
        _log_event(
            logging.INFO,
            "create_workbook completed",
            event="tool_call",
            tool_name="create_workbook",
            correlation_id=_current_correlation_id(),
        )
        return {
            "ok": True,
            "path": str(target),
            "relative_path": _relative_path(target),
            "sheets": wb.sheetnames,
            "backup_path": str(backup_path) if backup_path else None,
        }


@mcp.tool()
def list_workbooks() -> dict[str, Any]:
    """List workbook files under EXCEL_MCP_ROOT."""
    files = []
    for path in sorted(ROOT.rglob("*")):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        files.append(
            {
                "path": _relative_path(path),
                "size_bytes": path.stat().st_size,
                "modified_at": datetime.fromtimestamp(path.stat().st_mtime, UTC).isoformat(),
            }
        )
    return {"ok": True, "root": str(ROOT), "count": len(files), "files": files}


@mcp.tool()
def get_workbook_info(path: str) -> dict[str, Any]:
    """Return workbook metadata and per-sheet dimensions."""
    target = _safe_path(path)
    return _workbook_metadata(target)


@mcp.tool()
def list_sheets(path: str) -> dict[str, Any]:
    """Return workbook sheet names."""
    target = _safe_path(path)
    if not target.exists():
        raise FileNotFoundError(f"Workbook not found: {target}")

    wb = load_workbook(target, data_only=False, read_only=True)
    return {"ok": True, "path": str(target), "relative_path": _relative_path(target), "sheets": wb.sheetnames}


@mcp.tool()
def read_sheet(
    path: str,
    sheet: str,
    max_rows: int = 200,
    offset: int = 0,
    limit: int | None = None,
    columns: list[str] | None = None,
) -> dict[str, Any]:
    """Read a worksheet as records using the first row as headers."""
    if offset < 0:
        raise ValueError("offset must be >= 0")
    effective_limit = limit if limit is not None else max_rows
    if effective_limit <= 0:
        raise ValueError("limit must be >= 1")

    target = _safe_path(path)
    if not target.exists():
        raise FileNotFoundError(f"Workbook not found: {target}")

    rows, selected_columns, total_rows = _sheet_records(
        target,
        sheet,
        offset=offset,
        limit=effective_limit,
        columns=columns,
    )
    return {
        "ok": True,
        "path": str(target),
        "relative_path": _relative_path(target),
        "sheet": sheet,
        "rows": rows,
        "total_rows": total_rows,
        "returned_rows": len(rows),
        "offset": offset,
        "limit": effective_limit,
        "columns": selected_columns,
    }


@mcp.tool()
def read_range(path: str, sheet: str, start_cell: str, end_cell: str) -> dict[str, Any]:
    """Read a rectangular range and return values as a row matrix."""
    target = _safe_path(path)
    if not target.exists():
        raise FileNotFoundError(f"Workbook not found: {target}")

    min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
    wb = load_workbook(target, data_only=False, read_only=True)
    if sheet not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {sheet}")
    ws = wb[sheet]
    values = [
        [_normalize_cell_value(cell) for cell in row]
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    ]
    return {
        "ok": True,
        "path": str(target),
        "relative_path": _relative_path(target),
        "sheet": sheet,
        "range": f"{start_cell}:{end_cell}",
        "row_count": len(values),
        "column_count": (max_col - min_col) + 1,
        "values": values,
    }


@mcp.tool()
def append_rows(path: str, sheet: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Append record rows to a worksheet, creating headers when needed."""
    if not rows:
        raise ValueError("rows must not be empty")

    target = _safe_path(path)
    _ensure_parent(target)

    with _correlation_scope():
        with _locked_workbook(target):
            backup_path = _backup_workbook(target)
            if target.exists():
                wb = load_workbook(target)
            else:
                wb = Workbook()

            ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)

            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                default = wb["Sheet"]
                if (
                    default.max_row == 1
                    and default.max_column == 1
                    and default["A1"].value is None
                    and default.title != sheet
                ):
                    wb.remove(default)

            header_row = []
            if ws.max_row >= 1:
                header_row = [cell.value for cell in ws[1] if cell.value not in (None, "")]

            if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
                header_row = []

            incoming_headers = list(rows[0].keys())
            if not header_row:
                for idx, col in enumerate(incoming_headers, start=1):
                    ws.cell(row=1, column=idx, value=col)
                header_row = incoming_headers
            else:
                for col in incoming_headers:
                    if col not in header_row:
                        header_row.append(col)
                        ws.cell(row=1, column=len(header_row), value=col)

            start_row = ws.max_row + 1
            for row_index, row_obj in enumerate(rows):
                current_row = start_row + row_index
                for idx, col in enumerate(header_row, start=1):
                    ws.cell(
                        row=current_row,
                        column=idx,
                        value=_serializable_cell_value(row_obj.get(col)),
                    )

            wb.save(target)

        _log_event(
            logging.INFO,
            "append_rows completed",
            event="tool_call",
            tool_name="append_rows",
            correlation_id=_current_correlation_id(),
        )
        return {
            "ok": True,
            "path": str(target),
            "relative_path": _relative_path(target),
            "sheet": sheet,
            "appended": len(rows),
            "start_row": start_row,
            "backup_path": str(backup_path) if backup_path else None,
        }


@mcp.tool()
def write_cells(path: str, sheet: str, updates: list[dict[str, Any]]) -> dict[str, Any]:
    """Write explicit cell updates such as {'cell': 'B2', 'value': 123}."""
    if not updates:
        raise ValueError("updates must not be empty")

    target = _safe_path(path)
    _ensure_parent(target)

    with _correlation_scope():
        with _locked_workbook(target):
            backup_path = _backup_workbook(target)
            if target.exists():
                wb = load_workbook(target)
            else:
                wb = Workbook()

            ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)
            for item in updates:
                cell = str(item["cell"]).strip().upper()
                value = item.get("value")
                ws[cell] = _serializable_cell_value(value)

            wb.save(target)

        _log_event(
            logging.INFO,
            "write_cells completed",
            event="tool_call",
            tool_name="write_cells",
            correlation_id=_current_correlation_id(),
        )
        return {
            "ok": True,
            "path": str(target),
            "relative_path": _relative_path(target),
            "sheet": sheet,
            "updated_cells": len(updates),
            "backup_path": str(backup_path) if backup_path else None,
        }


@mcp.tool()
def write_range(path: str, sheet: str, start_cell: str, values: list[list[Any]]) -> dict[str, Any]:
    """Write a 2D matrix into a worksheet starting at start_cell."""
    if not values:
        raise ValueError("values must not be empty")
    row_lengths = {len(row) for row in values}
    if not row_lengths or 0 in row_lengths or len(row_lengths) != 1:
        raise ValueError("values must be a non-empty rectangular matrix")

    target = _safe_path(path)
    _ensure_parent(target)
    start_row, start_col = coordinate_to_tuple(start_cell)
    width = row_lengths.pop()
    end_row = start_row + len(values) - 1
    end_col = start_col + width - 1
    end_cell = f"{get_column_letter(end_col)}{end_row}"

    with _correlation_scope():
        with _locked_workbook(target):
            backup_path = _backup_workbook(target)
            if target.exists():
                wb = load_workbook(target)
            else:
                wb = Workbook()
            ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)

            for row_offset, row_values in enumerate(values):
                for col_offset, value in enumerate(row_values):
                    ws.cell(
                        row=start_row + row_offset,
                        column=start_col + col_offset,
                        value=_serializable_cell_value(value),
                    )
            wb.save(target)

        _log_event(
            logging.INFO,
            "write_range completed",
            event="tool_call",
            tool_name="write_range",
            correlation_id=_current_correlation_id(),
        )
        return {
            "ok": True,
            "path": str(target),
            "relative_path": _relative_path(target),
            "sheet": sheet,
            "range": f"{start_cell}:{end_cell}",
            "rows_written": len(values),
            "columns_written": width,
            "backup_path": str(backup_path) if backup_path else None,
        }


@mcp.tool()
def sheet_dimensions(path: str, sheet: str) -> dict[str, Any]:
    """Return row and column counts for a worksheet."""
    target = _safe_path(path)
    wb, ws = _get_sheet(target, sheet)
    return {
        "ok": True,
        "path": str(target),
        "relative_path": _relative_path(target),
        "sheet": sheet,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "last_column_letter": get_column_letter(ws.max_column) if ws.max_column else None,
        "sheets": wb.sheetnames,
    }


@mcp.tool()
def apply_hvdc_export(path: str, export: dict[str, Any], overwrite: bool = True) -> dict[str, Any]:
    """Create a workbook from the shared HVDC row-set export contract."""
    target = _safe_path(path)
    _ensure_parent(target)

    if str(export.get("version") or "").strip() != EXPORT_VERSION:
        raise ValueError(f"Unsupported export version: {export.get('version')}")
    sheets = export.get("sheets")
    if not isinstance(sheets, list):
        raise ValueError("export.sheets must be a list")
    sheet_rows = {
        str(item.get("name")): item.get("rows") or []
        for item in sheets
        if isinstance(item, dict) and item.get("name")
    }

    with _correlation_scope():
        with _locked_workbook(target):
            if target.exists() and not overwrite:
                raise FileExistsError(f"Workbook already exists: {target}")
            backup_path = _backup_workbook(target)
            wb = Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            for sheet_name in DEFAULT_EXPORT_SHEETS:
                ws = wb.create_sheet(title=sheet_name)
                rows = sheet_rows.get(sheet_name) or []
                if any(not isinstance(row, dict) for row in rows):
                    raise ValueError(f"Sheet rows must be dict objects: {sheet_name}")
                _write_row_dicts(ws, rows)
            wb.save(target)

        _log_event(
            logging.INFO,
            "apply_hvdc_export completed",
            event="tool_call",
            tool_name="apply_hvdc_export",
            correlation_id=_current_correlation_id(),
        )
        metadata = _workbook_metadata(target)
        return {
            **metadata,
            "export_kind": export.get("kind"),
            "workbook_name": export.get("workbook_name"),
            "backup_path": str(backup_path) if backup_path else None,
        }


if __name__ == "__main__":
    _log_event(
        logging.INFO,
        "Excel MCP starting",
        event="startup",
        path="/mcp",
        method="streamable-http",
        status_code=200,
        correlation_id=_new_correlation_id(),
    )
    mcp.run(transport="streamable-http")
